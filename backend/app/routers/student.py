"""
学员管理路由模块
包含学员信息、课程账户、操作（分班/转班/增减课时/转课时/退费/结业/停课）、头像/背景图、导入/导出、作品管理
"""
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Request, Body
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import or_, func, and_
from typing import Optional, List
from datetime import date, datetime, timedelta
from pydantic import BaseModel
import io
import openpyxl
import os
import uuid
import shutil
from urllib.parse import quote
import random
import logging

from app.database import get_db
from app.models import (
    Student, StudentCourse, Order, Class, Course, Attendance, IntegralRecord,
    EnrollmentHistory, Refund, Schedule, LearningCard, CardTransaction, ValidityLog,
    StudentWork
)
from app.schemas import StudentCreate, StudentUpdate
from app.utils import (
    generate_student_uuid, calculate_age, generate_order_no,
    deduct_from_orders_with_detail, calculate_payable_amount, calculate_refund_info
)
from app.services.card_service import CardService, get_card_details
from app.services.transaction_service import TransactionService
from app.services.student_operation_service import StudentOperationService
from app.services.image_service import ImageService
from dateutil.parser import parse as parse_date
from fastapi.responses import StreamingResponse
from PIL import Image as PILImage

logger = logging.getLogger(__name__)
router = APIRouter()

# ===================================================================
# 1. 请求模型定义
# ===================================================================

class TransferHoursRequest(BaseModel):
    from_course_id: int
    to_course_id: int
    transfer_hours: int
    reason: str = ''
    make_up_diff: bool = False
    pay_amount: float = 0
    payment_method: str = '微信'
    to_class_id: Optional[int] = None
    to_amount: float = 0
    to_hours: int = 0
    validity_days: Optional[int] = None


class DeleteWorksRequest(BaseModel):
    ids: List[int]


# ===================================================================
# 2. 辅助函数（私有方法）
# ===================================================================

def _build_contact_display(student: Student) -> tuple:
    """构建联系人显示字符串"""
    contacts = student.contacts if hasattr(student, 'contacts') and student.contacts else []
    primary = ''
    secondary = ''
    if contacts:
        if len(contacts) > 0:
            c = contacts[0]
            primary = f"{c.get('name','')}({c.get('relation','')}) {c.get('phone','')}"
        if len(contacts) > 1:
            c = contacts[1]
            secondary = f"{c.get('name','')}({c.get('relation','')}) {c.get('phone','')}"
    else:
        if student.primary_contact_name:
            primary = f"{student.primary_contact_name}({student.primary_contact_relation or ''}) {student.primary_contact_phone or ''}"
        if student.secondary_contact_name:
            secondary = f"{student.secondary_contact_name}({student.secondary_contact_relation or ''}) {student.secondary_contact_phone or ''}"
    return primary, secondary


def _format_card_item(card: LearningCard, student: Student, db: Session) -> dict:
    """格式化单个学习卡片的输出数据"""
    course = db.query(Course).get(card.course_id)
    class_obj = db.query(Class).get(card.class_id) if card.class_id else None

    # 统计订单数据
    orders = db.query(Order).filter(
        Order.student_id == student.id,
        Order.course_id == card.course_id,
        Order.is_active == True,
        Order.is_invalid == False,
        Order.enroll_type != '尾款补费'
    ).order_by(Order.created_at.asc()).all()
    arrears = sum(max(0, o.payable_amount - o.total_paid) for o in orders)

    total_purchased = db.query(func.sum(Order.purchase_hours)).filter(
        Order.student_id == student.id,
        Order.course_id == card.course_id,
        Order.is_active == True,
        Order.is_invalid == False,
        Order.enroll_type != '尾款补费'
    ).scalar() or 0

    total_gift = db.query(func.sum(Order.gift_hours)).filter(
        Order.student_id == student.id,
        Order.course_id == card.course_id,
        Order.is_active == True,
        Order.is_invalid == False,
        Order.enroll_type != '尾款补费'
    ).scalar() or 0

    total_purchased_amount = db.query(func.sum(Order.payable_amount)).filter(
        Order.student_id == student.id,
        Order.course_id == card.course_id,
        Order.is_active == True,
        Order.is_invalid == False,
        Order.enroll_type != '尾款补费'
    ).scalar() or 0.0

    # 考勤统计
    attendance_stats = db.query(
        func.sum(Attendance.deduct_hours).label('hours'),
        func.sum(Attendance.deduct_amount).label('amount')
    ).join(Schedule, Attendance.schedule_id == Schedule.id).join(Class, Schedule.class_id == Class.id).filter(
        Attendance.student_id == student.id,
        Class.course_id == card.course_id
    ).first()
    att_deduct_hours = round(float(attendance_stats.hours or 0), 2)
    att_deduct_amount = round(float(attendance_stats.amount or 0), 2)

    # 自定义流水聚合
    txn_agg = db.query(
        CardTransaction.change_type,
        func.sum(CardTransaction.amount_change).label('total_amount'),
        func.sum(CardTransaction.paid_hours_change + CardTransaction.gift_hours_change).label('total_hours')
    ).filter(
        CardTransaction.card_id == card.id,
        CardTransaction.change_type.in_(['custom_add', 'custom_sub', 'transfer_in', 'transfer_out'])
    ).group_by(CardTransaction.change_type).all()

    agg_map = {ct: (amt, hrs) for ct, amt, hrs in txn_agg}
    custom_add_amount = round(float(agg_map.get('custom_add', (0,0))[0]), 2)
    custom_sub_amount = abs(round(float(agg_map.get('custom_sub', (0,0))[0]), 2))
    transfer_in_amount = round(float(agg_map.get('transfer_in', (0,0))[0]), 2)
    transfer_out_amount = abs(round(float(agg_map.get('transfer_out', (0,0))[0]), 2))

    custom_add_hours = agg_map.get('custom_add', (0,0))[1] or 0
    custom_sub_hours = abs(agg_map.get('custom_sub', (0,0))[1] or 0)
    transfer_in_hours = agg_map.get('transfer_in', (0,0))[1] or 0
    transfer_out_hours = abs(agg_map.get('transfer_out', (0,0))[1] or 0)

    validity_display = '-'
    if card.validity_end_date:
        validity_display = card.validity_end_date.strftime('%Y-%m-%d')
    elif card.manual_validity_days:
        validity_display = (date.today() + timedelta(days=card.manual_validity_days)).strftime('%Y-%m-%d')

    total_leave_quota = card.total_leave_quota or 0
    used_leave = card.used_leave or 0
    leave_display = '不允许' if total_leave_quota == 0 else f"{used_leave}/{total_leave_quota}"

    return {
        "course_id": card.course_id,
        "course_name": course.name if course else '',
        "class_id": card.class_id,
        "class_name": class_obj.name if class_obj else ('未分班' if card.status == 'active' else '已退费'),
        "total_purchased": total_purchased,
        "remaining_hours": card.remaining_paid_hours or 0,
        "total_gift": total_gift,
        "remaining_gift": card.remaining_gift_hours or 0,
        "remaining_amount": card.remaining_amount or 0.0,
        "total_purchased_amount": total_purchased_amount,
        "validity_display": validity_display,
        "status": card.status,
        "suspended_start": card.suspended_start.strftime('%Y-%m-%d') if card.suspended_start else None,
        "suspended_end": card.suspended_end.strftime('%Y-%m-%d') if card.suspended_end else None,
        "total_deducted": att_deduct_hours,
        "total_custom_add": custom_add_hours,
        "total_custom_sub": custom_sub_hours,
        "total_transfer_in": transfer_in_hours,
        "total_transfer_out": transfer_out_hours,
        "attendance_deducted_hours": att_deduct_hours,
        "attendance_deducted_amount": att_deduct_amount,
        "custom_add_amount": custom_add_amount,
        "custom_sub_amount": custom_sub_amount,
        "transfer_in_amount": transfer_in_amount,
        "transfer_out_amount": transfer_out_amount,
        "arrears": round(arrears, 2),
        "leave_display": leave_display,
        "orders": [{
            "order_no": o.order_no,
            "enroll_type": o.enroll_type,
            "purchase_hours": o.purchase_hours,
            "total_price": o.total_price,
            "discount_amount": o.discount_amount,
            "payable_amount": o.payable_amount,
            "paid_amount": o.paid_amount,
            "total_paid": o.total_paid,
            "remaining_hours": 0,
            "remaining_amount": 0.0,
            "gift_hours": o.gift_hours,
            "gift_remaining": 0,
            "created_at": str(o.created_at),
            "actual_unit_price": round(o.payable_amount / o.purchase_hours, 2) if o.purchase_hours > 0 else 0
        } for o in orders],
        "total_leave_quota": card.total_leave_quota,
        "used_leave": card.used_leave
    }


# ===================================================================
# 3. 学员列表与搜索
# ===================================================================

@router.get("/student/students")
def student_list(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query('active'),
    class_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取在学学员列表（分页）"""
    query = db.query(Student).filter(Student.is_archived == False)

    if search:
        query = query.filter(
            or_(Student.name.contains(search), Student.phone.contains(search))
        )

    total = query.count()
    students = query.order_by(Student.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    if not students:
        return {"code": 0, "data": {"items": [], "total": 0}}

    student_ids = [s.id for s in students]
    cards_query = db.query(LearningCard).filter(
        LearningCard.student_id.in_(student_ids),
        LearningCard.status.in_(['active', 'suspended'])
    ).options(
        selectinload(LearningCard.course),
        selectinload(LearningCard.class_),
        selectinload(LearningCard.validity_logs)
    )

    cards_by_student = {}
    for card in cards_query:
        cards_by_student.setdefault(card.student_id, []).append(card)

    all_course_ids = {card.course_id for cards in cards_by_student.values() for card in cards}
    all_class_ids = {card.class_id for cards in cards_by_student.values() for card in cards if card.class_id}

    courses = {c.id: c for c in db.query(Course).filter(Course.id.in_(all_course_ids)).all()}
    classes = {c.id: c for c in db.query(Class).filter(Class.id.in_(all_class_ids)).all()}

    order_stats = db.query(
        Order.student_id,
        Order.course_id,
        func.sum(Order.purchase_hours).label('total_purchased'),
        func.sum(Order.gift_hours).label('total_gift'),
        func.sum(Order.payable_amount).label('total_amount'),
        func.sum(func.abs(Order.payable_amount - Order.total_paid)).label('total_arrears')
    ).filter(
        Order.student_id.in_(student_ids),
        Order.course_id.in_(all_course_ids),
        Order.is_active == True,
        Order.is_invalid == False,
        Order.enroll_type.notin_(['尾款补费', '转课时'])
    ).group_by(Order.student_id, Order.course_id).all()

    order_stats_dict = {}
    for stat in order_stats:
        key = (stat.student_id, stat.course_id)
        order_stats_dict[key] = {
            'total_purchased': stat.total_purchased or 0,
            'total_gift': stat.total_gift or 0,
            'total_amount': stat.total_amount or 0,
            'total_arrears': stat.total_arrears or 0
        }

    result = []
    for student in students:
        cards = cards_by_student.get(student.id, [])
        if not cards:
            continue

        primary, secondary = _build_contact_display(student)

        for card in cards:
            course = courses.get(card.course_id)
            class_obj = classes.get(card.class_id) if card.class_id else None

            stats_key = (student.id, card.course_id)
            stats = order_stats_dict.get(stats_key, {})
            total_purchased = stats.get('total_purchased', 0)
            total_gift = stats.get('total_gift', 0)
            arrears = stats.get('total_arrears', 0)

            remaining_paid = card.remaining_paid_hours or 0
            remaining_gift = card.remaining_gift_hours or 0
            remaining_amount = card.remaining_amount or 0.0

            total_leave_quota = card.total_leave_quota or 0
            used_leave = card.used_leave or 0
            if total_leave_quota == 0:
                leave_display = '不允许'
                remaining_leaves = 0
            else:
                leave_display = f"{used_leave}/{total_leave_quota}"
                remaining_leaves = max(0, total_leave_quota - used_leave)

            is_suspended = card.status == 'suspended'

            validity_display = '-'
            if card.validity_end_date:
                validity_display = card.validity_end_date.strftime('%Y-%m-%d')
            elif card.manual_validity_days:
                validity_display = (date.today() + timedelta(days=card.manual_validity_days)).strftime('%Y-%m-%d')

            result.append({
                "id": student.id,
                "name": student.name,
                "phone": student.phone,
                "gender": student.gender,
                "age": student.age,
                "total_integral": student.total_integral,
                "primary_contact": primary,
                "secondary_contact": secondary,
                "course_name": course.name if course else '',
                "class_name": class_obj.name if class_obj else '未分班',
                "class_id": card.class_id,
                "course_id": card.course_id,
                "remaining_hours": remaining_paid,
                "remaining_amount": remaining_amount,
                "remaining_gift": remaining_gift,
                "total_purchased": total_purchased,
                "total_gift": total_gift,
                "arrears": round(arrears, 2),
                "status": card.status,
                "is_suspended": is_suspended,
                "suspended_start": card.suspended_start.strftime('%Y-%m-%d') if card.suspended_start else None,
                "suspended_end": card.suspended_end.strftime('%Y-%m-%d') if card.suspended_end else None,
                "validity_display": validity_display,
                "leave_display": leave_display,
                "remaining_leaves": remaining_leaves,
                "total_leave_quota": card.total_leave_quota,
                "used_leave": card.used_leave,
                "avatar": student.avatar or ""
            })

    return {"code": 0, "data": {"items": result, "total": total}}


@router.get("/student/graduated-students")
def graduated_students(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取已结业/归档学员列表"""
    query = db.query(Student).filter(Student.is_archived == True)
    if search:
        query = query.filter(or_(Student.name.contains(search), Student.phone.contains(search)))
    total = query.count()
    students = query.order_by(Student.updated_at.desc()).offset((page-1)*page_size).limit(page_size).all()
    result = [{
        "id": s.id,
        "name": s.name,
        "phone": s.phone,
        "gender": s.gender,
        "age": s.age,
        "total_integral": s.total_integral,
        "avatar": ImageService.get_url(s.avatar)
    } for s in students]
    return {"code": 0, "data": {"items": result, "total": total}}


@router.get("/student/search")
def search_students(
    keyword: str = Query(''),
    db: Session = Depends(get_db)
):
    """搜索学员（用于下拉选择）"""
    if not keyword:
        return {"code": 0, "data": []}
    students = db.query(Student).filter(
        or_(Student.name.contains(keyword), Student.phone.contains(keyword))
    ).limit(20).all()
    result = [{
        "id": s.id,
        "name": s.name,
        "phone": s.phone,
        "gender": s.gender,
        "birthday": str(s.birthday) if s.birthday else None,
        "age": s.age,
        "introducer": s.introducer,
        "note": s.note,
        "primary_contact_name": s.primary_contact_name,
        "primary_contact_relation": s.primary_contact_relation,
        "primary_contact_phone": s.primary_contact_phone,
        "secondary_contact_name": s.secondary_contact_name,
        "secondary_contact_relation": s.secondary_contact_relation,
        "secondary_contact_phone": s.secondary_contact_phone,
        "contacts": s.contacts or [],
        "avatar": ImageService.get_url(s.avatar)
    } for s in students]
    return {"code": 0, "data": result}


# ===================================================================
# 4. 学员详情与课程账户
# ===================================================================

@router.get("/student/student/{student_id}")
def student_detail(student_id: int, db: Session = Depends(get_db)):
    """获取学员详细信息"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    return {
        "code": 0,
        "data": {
            "id": student.id,
            "student_uuid": student.student_uuid,
            "name": student.name,
            "phone": student.phone,
            "gender": student.gender,
            "birthday": str(student.birthday) if student.birthday else None,
            "age": student.age,
            "introducer": student.introducer,
            "note": student.note,
            "total_integral": student.total_integral,
            "avatar": ImageService.get_url(student.avatar),
            "card_background": ImageService.get_url(student.card_background),
            "primary_contact_name": student.primary_contact_name,
            "primary_contact_relation": student.primary_contact_relation,
            "primary_contact_phone": student.primary_contact_phone,
            "secondary_contact_name": student.secondary_contact_name,
            "secondary_contact_relation": student.secondary_contact_relation,
            "secondary_contact_phone": student.secondary_contact_phone,
            "contacts": student.contacts or [],
            # 新增字段
            "nation": student.nation or '',
            "id_card": student.id_card or '',
            "school": student.school or ''
        }
    }


@router.get("/student/student/{student_id}/courses")
def student_courses_detail(student_id: int, db: Session = Depends(get_db)):
    """获取学员课程账户详情"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    cards = db.query(LearningCard).filter(LearningCard.student_id == student_id).all()
    result = [_format_card_item(card, student, db) for card in cards]
    return {"code": 0, "data": result}


@router.post("/student/students")
def create_student(data: StudentCreate, db: Session = Depends(get_db)):
    """创建新学员（如果手机号已存在则返回已有学员）"""
    existing = db.query(Student).filter(Student.phone == data.phone).first()
    if existing:
        return {"code": 0, "data": {"id": existing.id, "name": existing.name, "phone": existing.phone}}
    birth_date = parse_date(data.birth_date).date() if data.birth_date else None
    age = calculate_age(birth_date) if birth_date else None
    student = Student(
        student_uuid=generate_student_uuid(),
        name=data.name,
        phone=data.phone,
        gender=data.gender or '',
        birthday=birth_date,
        age=age,
        introducer=data.introducer or '',
        note=data.note or '',
        primary_contact_name=data.primary_contact_name or '',
        primary_contact_relation=data.primary_contact_relation or '',
        primary_contact_phone=data.primary_contact_phone or '',
        secondary_contact_name=data.secondary_contact_name or '',
        secondary_contact_relation=data.secondary_contact_relation or '',
        secondary_contact_phone=data.secondary_contact_phone or '',
        contacts=data.contacts if data.contacts else []
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return {"code": 0, "data": {"id": student.id, "name": student.name, "phone": student.phone}}


@router.put("/student/student/{student_id}")
def update_student(student_id: int, data: StudentUpdate, db: Session = Depends(get_db)):
    """更新学员信息（含新增字段 nation, id_card, school）"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    update_data = data.model_dump(exclude_unset=True)

    if 'birth_date' in update_data and update_data['birth_date']:
        birth_date = parse_date(update_data['birth_date']).date()
        student.birthday = birth_date
        student.age = calculate_age(birth_date)
        del update_data['birth_date']

    if 'contacts' in update_data:
        student.contacts = update_data['contacts']
        from sqlalchemy.orm import attributes
        attributes.flag_modified(student, 'contacts')
        del update_data['contacts']

    # 处理 card_background 和新增字段
    for field in ['card_background', 'nation', 'id_card', 'school']:
        if field in update_data:
            setattr(student, field, update_data[field])
            del update_data[field]

    for field, value in update_data.items():
        if hasattr(student, field):
            setattr(student, field, value)

    db.commit()
    return {"code": 0, "message": "更新成功"}


# ===================================================================
# 5. 头像与卡片背景图管理
# ===================================================================

@router.post("/student/student/{student_id}/avatar")
async def upload_avatar(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """上传学员头像"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    if student.avatar:
        ImageService.delete(student.avatar)
    relative_path = await ImageService.save(file, "avatar")
    student.avatar = relative_path
    db.commit()
    full_url = ImageService.get_url(relative_path)
    return {"code": 0, "data": {"avatar_url": full_url}, "message": "上传成功"}


@router.delete("/student/student/{student_id}/avatar")
def delete_avatar(student_id: int, db: Session = Depends(get_db)):
    """删除学员头像"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    if student.avatar:
        ImageService.delete(student.avatar)
        student.avatar = None
        db.commit()
    return {"code": 0, "message": "头像已删除"}


@router.post("/student/student/{student_id}/card-background")
async def upload_card_background(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """上传学员卡片背景图"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    if student.card_background:
        ImageService.delete(student.card_background)
    relative_path = await ImageService.save(file, "card_bg")
    student.card_background = relative_path
    db.commit()
    full_url = ImageService.get_url(relative_path)
    return {"code": 0, "data": {"card_background_url": full_url}, "message": "上传成功"}


@router.delete("/student/student/{student_id}/card-background")
def delete_card_background(student_id: int, db: Session = Depends(get_db)):
    """删除学员卡片背景图"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")
    if student.card_background:
        ImageService.delete(student.card_background)
        student.card_background = None
        db.commit()
    return {"code": 0, "message": "背景图已删除"}


# ===================================================================
# 6. 学员操作（分班/转班/退班/增减课时/转课时/退费/结业/停课）
# ===================================================================

@router.post("/student/student/{student_id}/assign-class")
def assign_class(
    student_id: int,
    course_id: int = Query(...),
    class_id: int = Query(...),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.assign_class(student_id, course_id, class_id)
    db.commit()
    return {"code": 0, "message": "分班成功"}


@router.post("/student/student/{student_id}/transfer-class")
def transfer_class(
    student_id: int,
    course_id: int = Query(...),
    to_class_id: int = Query(...),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.transfer_class(student_id, course_id, to_class_id)
    db.commit()
    return {"code": 0, "message": "转班成功"}


@router.post("/student/student/{student_id}/drop-class")
def drop_class(
    student_id: int,
    course_id: int = Query(...),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.drop_class(student_id, course_id)
    db.commit()
    return {"code": 0, "message": "退班成功"}


@router.post("/student/student/{student_id}/adjust-hours")
def adjust_hours(
    student_id: int,
    course_id: int = Query(...),
    change_hours: int = Query(...),
    reason: str = Query(''),
    occurrence_date: Optional[str] = Query(None),
    performance: bool = Query(False),
    teacher_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    occ_date = parse_date(occurrence_date).date() if occurrence_date else date.today()
    service = StudentOperationService(db)
    service.adjust_hours(student_id, course_id, change_hours, reason, occ_date, performance, teacher_id)
    db.commit()
    return {"code": 0, "message": "操作成功"}


@router.post("/student/student/{student_id}/gift-hours")
def gift_hours(
    student_id: int,
    course_id: int = Query(...),
    change_hours: int = Query(...),
    reason: str = Query(''),
    occurrence_date: Optional[str] = Query(None),
    performance: bool = Query(False),
    teacher_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    occ_date = parse_date(occurrence_date).date() if occurrence_date else date.today()
    service = StudentOperationService(db)
    service.adjust_gift_hours(student_id, course_id, change_hours, reason, occ_date, performance, teacher_id)
    db.commit()
    return {"code": 0, "message": "操作成功"}


@router.post("/student/student/{student_id}/transfer-hours")
def transfer_hours(
    student_id: int,
    data: TransferHoursRequest,
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.transfer_hours(
        student_id=student_id,
        from_course_id=data.from_course_id,
        to_course_id=data.to_course_id,
        transfer_hours=data.transfer_hours,
        reason=data.reason,
        to_hours=data.to_hours,
        to_amount=data.to_amount,
        make_up_diff=data.make_up_diff,
        pay_amount=data.pay_amount,
        payment_method=data.payment_method,
        to_class_id=data.to_class_id,
        validity_days=data.validity_days
    )
    db.commit()
    return {"code": 0, "message": "转课时成功"}


@router.get("/student/student/{student_id}/transfer-preview")
def transfer_preview(
    student_id: int,
    course_id: int = Query(...),
    transfer_hours: int = Query(...),
    db: Session = Depends(get_db)
):
    """转课时金额预览"""
    try:
        gift, paid, amount = deduct_from_orders_with_detail(
            db, student_id, course_id, transfer_hours, deduct_gift=False
        )
        return {"code": 0, "data": {"amount": amount, "paid_hours": paid, "gift_hours": gift}}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, str(e))


@router.get("/student/student/{student_id}/refund-preview")
def refund_preview(
    student_id: int,
    course_id: int = Query(...),
    refund_hours: float = Query(...),
    db: Session = Depends(get_db)
):
    """退费预览"""
    tx_service = TransactionService(db)
    tx_service._refresh_card(student_id, course_id)
    try:
        info = calculate_refund_info(db, student_id, course_id, refund_hours)
        return {"code": 0, "data": info}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, str(e))


@router.post("/student/student/{student_id}/refund")
def refund_course(
    student_id: int,
    course_id: int = Query(...),
    refund_hours: float = Query(...),
    reason: str = Query(''),
    refund_method: str = Query('微信'),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.refund(student_id, course_id, refund_hours, reason, refund_method)
    db.commit()
    return {"code": 0, "message": "退费成功"}


@router.post("/student/student/{student_id}/graduate")
def graduate_student(
    student_id: int,
    course_id: int = Query(...),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.graduate(student_id, course_id)
    db.commit()
    return {"code": 0, "message": "结业成功"}


@router.post("/student/student/{student_id}/suspend")
def set_suspend(
    student_id: int,
    course_id: int = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...),
    reason: str = Query(''),
    db: Session = Depends(get_db)
):
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "日期格式错误")
    logger.info(f"停课请求: student={student_id}, course={course_id}, start={start_date}, end={end_date}")
    service = StudentOperationService(db)
    service.suspend(student_id, course_id, start, end, reason)
    db.commit()
    return {"code": 0, "message": "停课设置成功"}


@router.post("/student/student/{student_id}/cancel-suspend")
def cancel_suspend(
    student_id: int,
    course_id: int = Query(...),
    db: Session = Depends(get_db)
):
    service = StudentOperationService(db)
    service.cancel_suspend(student_id, course_id)
    db.commit()
    return {"code": 0, "message": "已取消停课"}


@router.post("/student/student/{student_id}/extend-validity")
def extend_validity(
    student_id: int,
    course_id: int = Query(...),
    extend_days: int = Query(..., ge=1, le=365),
    reason: str = Query(''),
    db: Session = Depends(get_db)
):
    """手动延长课程有效期"""
    card = db.query(LearningCard).filter(
        LearningCard.student_id == student_id,
        LearningCard.course_id == course_id
    ).first()
    if not card:
        raise HTTPException(404, "课程账户不存在")
    tx_service = TransactionService(db)
    tx_service.extend_validity(student_id, course_id, extend_days, reason)
    db.commit()
    db.refresh(card)
    return {
        "code": 0,
        "message": f"已延长有效期 {extend_days} 天",
        "data": {"validity_end_date": card.validity_end_date.strftime('%Y-%m-%d') if card.validity_end_date else None}
    }


# ===================================================================
# 7. 学员导出与导入
# ===================================================================

@router.get("/student/students/export")
def export_students(db: Session = Depends(get_db)):
    """导出在学学员列表为 Excel"""
    students = db.query(Student).filter(Student.is_archived == False).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学员列表"

    headers = [
        "学员姓名", "性别", "年龄", "手机号", "主要联系人", "备用联系人", "总积分",
        "报名课程", "所在班级", "购买课时", "剩余课时", "剩余金额", "赠送课时", "剩余赠送课时",
        "课时有效期", "请假次数", "状态", "停课起止"
    ]
    ws.append(headers)

    for student in students:
        primary, secondary = _build_contact_display(student)
        cards = db.query(LearningCard).filter(
            LearningCard.student_id == student.id,
            LearningCard.status.in_(['active', 'suspended'])
        ).all()

        if not cards:
            ws.append([
                student.name, student.gender or '', student.age or '', student.phone,
                primary, secondary, student.total_integral or 0,
                '-', '-', 0, 0, 0, 0, 0, '-', '-', '无课程', ''
            ])
        else:
            for card in cards:
                course = db.query(Course).get(card.course_id)
                class_obj = db.query(Class).get(card.class_id) if card.class_id else None

                total_purchased = db.query(func.sum(Order.purchase_hours)).filter(
                    Order.student_id == student.id,
                    Order.course_id == card.course_id,
                    Order.is_active == True,
                    Order.is_invalid == False,
                    Order.enroll_type.notin_(['尾款补费', '转课时'])
                ).scalar() or 0

                total_gift = db.query(func.sum(Order.gift_hours)).filter(
                    Order.student_id == student.id,
                    Order.course_id == card.course_id,
                    Order.is_active == True,
                    Order.is_invalid == False,
                    Order.enroll_type.notin_(['尾款补费', '转课时'])
                ).scalar() or 0

                validity_display = '-'
                if card.validity_end_date:
                    validity_display = card.validity_end_date.strftime('%Y-%m-%d')
                elif card.manual_validity_days:
                    validity_display = (date.today() + timedelta(days=card.manual_validity_days)).strftime('%Y-%m-%d')

                total_leave_quota = card.total_leave_quota or 0
                used_leave = card.used_leave or 0
                leave_display = '不允许' if total_leave_quota == 0 else f"{used_leave}/{total_leave_quota}"

                suspend_display = ''
                if card.status == 'suspended' and card.suspended_start and card.suspended_end:
                    suspend_display = f"{card.suspended_start.strftime('%Y-%m-%d')} 至 {card.suspended_end.strftime('%Y-%m-%d')}"
                elif card.status == 'suspended':
                    suspend_display = '停课中'
                else:
                    suspend_display = '正常'

                status_display = '停课' if card.status == 'suspended' else '在学'

                ws.append([
                    student.name, student.gender or '', student.age or '', student.phone,
                    primary, secondary, student.total_integral or 0,
                    course.name if course else '-', class_obj.name if class_obj else '未分班',
                    total_purchased,
                    card.remaining_paid_hours or 0,
                    card.remaining_amount or 0,
                    total_gift,
                    card.remaining_gift_hours or 0,
                    validity_display,
                    leave_display,
                    status_display,
                    suspend_display
                ])

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )


@router.post("/student/students/import")
async def import_students(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入学员（Excel）"""
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value.strip()] = idx

    required_fields = ['姓名', '手机号', '报名课程', '购买课时', '实付金额']
    for field in required_fields:
        if field not in headers:
            return {"code": 1, "message": f"模板缺少必填字段：{field}"}

    imported_count = 0
    errors = []
    timestamp_prefix = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[headers.get('姓名', 0)-1]:
            continue

        try:
            name = str(row[headers['姓名']-1]).strip() if headers.get('姓名') else ''
            gender = str(row[headers['性别']-1]).strip() if headers.get('性别') and row[headers['性别']-1] else ''
            phone = str(row[headers['手机号']-1]).strip() if headers.get('手机号') else ''
            birth_date_str = str(row[headers['出生日期']-1]).strip() if headers.get('出生日期') and row[headers['出生日期']-1] else ''
            introducer = str(row[headers['介绍人']-1]).strip() if headers.get('介绍人') and row[headers['介绍人']-1] else ''
            current_integral = int(row[headers['当前积分']-1]) if headers.get('当前积分') and row[headers['当前积分']-1] else 0

            primary_name = str(row[headers['主要联系人姓名']-1]).strip() if headers.get('主要联系人姓名') and row[headers['主要联系人姓名']-1] else ''
            primary_relation = str(row[headers['主要联系人关系']-1]).strip() if headers.get('主要联系人关系') and row[headers['主要联系人关系']-1] else ''
            primary_phone = str(row[headers['主要联系人电话']-1]).strip() if headers.get('主要联系人电话') and row[headers['主要联系人电话']-1] else ''
            secondary_name = str(row[headers['备用联系人姓名']-1]).strip() if headers.get('备用联系人姓名') and row[headers['备用联系人姓名']-1] else ''
            secondary_relation = str(row[headers['备用联系人关系']-1]).strip() if headers.get('备用联系人关系') and row[headers['备用联系人关系']-1] else ''
            secondary_phone = str(row[headers['备用联系人电话']-1]).strip() if headers.get('备用联系人电话') and row[headers['备用联系人电话']-1] else ''

            if not name or not phone:
                errors.append(f"第{row_idx}行：姓名或手机号为空")
                continue

            existing_student = db.query(Student).filter(Student.phone == phone).first()
            if existing_student:
                student = existing_student
                if primary_name:
                    student.primary_contact_name = primary_name
                    student.primary_contact_relation = primary_relation
                    student.primary_contact_phone = primary_phone
                if secondary_name:
                    student.secondary_contact_name = secondary_name
                    student.secondary_contact_relation = secondary_relation
                    student.secondary_contact_phone = secondary_phone
                student.total_integral = current_integral
            else:
                birth_date = None
                age = None
                if birth_date_str:
                    try:
                        birth_date = parse_date(birth_date_str).date()
                        age = calculate_age(birth_date)
                    except:
                        pass
                student = Student(
                    student_uuid=generate_student_uuid(),
                    name=name,
                    phone=phone,
                    gender=gender,
                    birthday=birth_date,
                    age=age,
                    introducer=introducer,
                    total_integral=current_integral,
                    primary_contact_name=primary_name,
                    primary_contact_relation=primary_relation,
                    primary_contact_phone=primary_phone,
                    secondary_contact_name=secondary_name,
                    secondary_contact_relation=secondary_relation,
                    secondary_contact_phone=secondary_phone,
                    is_archived=False
                )
                db.add(student)
                db.flush()
                if current_integral > 0:
                    integral_record = IntegralRecord(
                        student_id=student.id,
                        change_amount=current_integral,
                        reason="导入学员初始积分"
                    )
                    db.add(integral_record)

            course_name = str(row[headers['报名课程']-1]).strip() if headers.get('报名课程') else ''
            if not course_name:
                continue

            course = db.query(Course).filter(Course.name == course_name).first()
            if not course:
                errors.append(f"第{row_idx}行：课程【{course_name}】不存在")
                continue

            class_name = str(row[headers['班级名称']-1]).strip() if headers.get('班级名称') and row[headers['班级名称']-1] else ''
            class_id = None
            if class_name:
                class_obj = db.query(Class).filter(Class.name == class_name, Class.course_id == course.id).first()
                if class_obj:
                    class_id = class_obj.id
                else:
                    errors.append(f"第{row_idx}行：班级【{class_name}】不存在于课程【{course_name}】下")

            purchase_hours = int(row[headers['购买课时']-1]) if headers.get('购买课时') else 0
            if purchase_hours <= 0:
                errors.append(f"第{row_idx}行：购买课时必须大于0")
                continue

            gift_hours = int(row[headers['赠送课时']-1]) if headers.get('赠送课时') and row[headers['赠送课时']-1] else 0
            validity_days = int(row[headers['有效期天数']-1]) if headers.get('有效期天数') and row[headers['有效期天数']-1] else None
            leave_limit_value = str(row[headers['请假限制次数']-1]).strip() if headers.get('请假限制次数') and row[headers['请假限制次数']-1] else '不限制'

            if leave_limit_value.isdigit():
                leave_limit = '限制次数'
                leave_limit_count = int(leave_limit_value)
            elif leave_limit_value == '不允许':
                leave_limit = '不允许'
                leave_limit_count = 0
            else:
                leave_limit = '不限制'
                leave_limit_count = 0

            enroll_type = str(row[headers['报名类型']-1]).strip() if headers.get('报名类型') and row[headers['报名类型']-1] else '新报'
            if enroll_type not in ['新报', '续报', '扩科', '尾款补费']:
                enroll_type = '新报'

            discount_mode_raw = str(row[headers['优惠方式']-1]).strip() if headers.get('优惠方式') and row[headers['优惠方式']-1] else '无'
            discount_value = float(row[headers['优惠值']-1]) if headers.get('优惠值') and row[headers['优惠值']-1] else 0

            discount_mode = 'none'
            discount_rate = 0.0
            direct_reduction = 0.0
            if discount_mode_raw == '折扣':
                discount_mode = 'discount'
                discount_rate = discount_value
            elif discount_mode_raw == '直减':
                discount_mode = 'direct'
                direct_reduction = discount_value
            elif discount_mode_raw == '先折扣再直减':
                discount_mode = 'discount_then_direct'
                discount_rate = discount_value
            elif discount_mode_raw == '先直减再折扣':
                discount_mode = 'direct_then_discount'
                direct_reduction = discount_value
            else:
                discount_mode = 'none'

            payment_method = str(row[headers['支付方式']-1]).strip() if headers.get('支付方式') and row[headers['支付方式']-1] else '微信'
            total_paid = float(row[headers['实付金额']-1]) if headers.get('实付金额') and row[headers['实付金额']-1] else 0

            unit_price = course.unit_price
            total_price = purchase_hours * unit_price
            payable, discount = calculate_payable_amount(total_price, discount_mode, discount_rate, direct_reduction)
            if total_paid > 0:
                payable = total_paid
                discount = total_price - payable

            order_no = f"{timestamp_prefix}{imported_count+1:04d}{random.randint(10, 99)}"
            validity_type = 'days' if validity_days else None
            validity_value = str(validity_days) if validity_days else None

            order = Order(
                order_no=order_no,
                student_id=student.id,
                course_id=course.id,
                class_id=class_id,
                enroll_type=enroll_type,
                purchase_hours=purchase_hours,
                unit_price=unit_price,
                total_price=total_price,
                discount_mode=discount_mode,
                discount_rate=discount_rate,
                direct_reduction=direct_reduction,
                discount_amount=discount,
                payable_amount=payable,
                actual_unit_price=round(payable / purchase_hours, 2) if purchase_hours > 0 else 0,
                paid_amount=total_paid,
                total_paid=total_paid,
                payment_method=payment_method,
                gift_hours=gift_hours,
                validity_type=validity_type,
                validity_value=validity_value,
                leave_limit=leave_limit,
                leave_limit_count=leave_limit_count,
                is_active=True,
                is_invalid=False
            )
            db.add(order)
            db.flush()

            tx_service = TransactionService(db)
            tx_service.record_purchase(order, gift_hours)

            imported_count += 1

        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")
            db.rollback()

    if errors:
        db.rollback()
        return {"code": 1, "message": f"导入失败，共{len(errors)}条错误：{'；'.join(errors[:3])}"}
    else:
        db.commit()
        return {"code": 0, "message": f"成功导入 {imported_count} 条记录"}


@router.get("/student/students/import-template")
def download_import_template():
    """下载导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学员导入模板"

    headers = [
        "姓名", "性别", "手机号", "出生日期", "介绍人", "当前积分",
        "主要联系人姓名", "主要联系人关系", "主要联系人电话",
        "备用联系人姓名", "备用联系人关系", "备用联系人电话",
        "报名课程", "购买课时", "赠送课时", "有效期天数", "请假限制次数",
        "报名类型", "优惠方式", "优惠值", "班级名称",
        "报名时间", "支付方式", "实付金额"
    ]
    ws.append(headers)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "学员导入模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ===================================================================
# 8. 聚合学员列表（用于卡片视图）
# ===================================================================

@router.get("/students-aggregated")
def get_aggregated_students(
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(36, ge=1, le=200),
    student_name: Optional[str] = Query(None),
    student_phone: Optional[str] = Query(None),
    birth_month: Optional[int] = Query(None, ge=1, le=12),
    age_min: Optional[int] = Query(None),
    age_max: Optional[int] = Query(None),
    is_archived: bool = Query(False),
    course_name: Optional[str] = Query(None),
    class_name: Optional[str] = Query(None),
    teacher_name: Optional[str] = Query(None),
    has_class: Optional[bool] = Query(None),
    is_suspended: Optional[bool] = Query(None),
    remaining_hours_min: Optional[int] = Query(None),
    remaining_hours_max: Optional[int] = Query(None),
    validity_days_min: Optional[int] = Query(None),
    validity_days_max: Optional[int] = Query(None),
    leave_count_min: Optional[int] = Query(None),
    leave_count_max: Optional[int] = Query(None)
):
    """按学员聚合返回数据（支持学员级和课程级筛选）"""
    # 学员级筛选
    student_query = db.query(Student).filter(Student.is_archived == is_archived)
    if student_name:
        student_query = student_query.filter(Student.name.contains(student_name))
    if student_phone:
        student_query = student_query.filter(Student.phone.contains(student_phone))
    if birth_month:
        student_query = student_query.filter(func.extract('month', Student.birthday) == birth_month)
    if age_min is not None or age_max is not None:
        today = date.today()
        if age_min is not None:
            max_birth = today.replace(year=today.year - age_min)
            student_query = student_query.filter(Student.birthday <= max_birth)
        if age_max is not None:
            min_birth = today.replace(year=today.year - age_max - 1)
            student_query = student_query.filter(Student.birthday >= min_birth)
    student_ids = [s.id for s in student_query.all()]
    if not student_ids:
        return {"code": 0, "data": {"items": [], "total": 0}}

    # 课程级筛选
    card_query = db.query(LearningCard).filter(
        LearningCard.student_id.in_(student_ids),
        LearningCard.status.in_(['active', 'suspended'])
    )
    if course_name:
        card_query = card_query.join(Course, LearningCard.course_id == Course.id).filter(Course.name.contains(course_name))
    if class_name:
        card_query = card_query.join(Class, LearningCard.class_id == Class.id).filter(Class.name.contains(class_name))
    if teacher_name:
        card_query = card_query.join(Class, LearningCard.class_id == Class.id).join(Teacher, Class.teacher_id == Teacher.id).filter(Teacher.name.contains(teacher_name))
    if has_class is not None:
        if has_class:
            card_query = card_query.filter(LearningCard.class_id.isnot(None))
        else:
            card_query = card_query.filter(LearningCard.class_id.is_(None))
    if is_suspended is not None:
        if is_suspended:
            card_query = card_query.filter(LearningCard.status == 'suspended')
        else:
            card_query = card_query.filter(LearningCard.status == 'active')
    if remaining_hours_min is not None:
        card_query = card_query.filter(LearningCard.remaining_paid_hours >= remaining_hours_min)
    if remaining_hours_max is not None:
        card_query = card_query.filter(LearningCard.remaining_paid_hours <= remaining_hours_max)
    if validity_days_min is not None or validity_days_max is not None:
        today = date.today()
        if validity_days_min is not None:
            card_query = card_query.filter(LearningCard.validity_end_date >= today + timedelta(days=validity_days_min))
        if validity_days_max is not None:
            card_query = card_query.filter(LearningCard.validity_end_date <= today + timedelta(days=validity_days_max))
    if leave_count_min is not None or leave_count_max is not None:
        if leave_count_min is not None:
            card_query = card_query.filter(LearningCard.used_leave >= leave_count_min)
        if leave_count_max is not None:
            card_query = card_query.filter(LearningCard.used_leave <= leave_count_max)

    cards = card_query.all()
    student_cards_map = {}
    for card in cards:
        student_cards_map.setdefault(card.student_id, []).append(card)

    filtered_student_ids = list(student_cards_map.keys())
    if not filtered_student_ids:
        return {"code": 0, "data": {"items": [], "total": 0}}

    total = len(filtered_student_ids)
    offset = (page - 1) * page_size
    paged_student_ids = filtered_student_ids[offset:offset + page_size]

    students = db.query(Student).filter(Student.id.in_(paged_student_ids)).all()
    students_dict = {s.id: s for s in students}

    result = []
    for sid in paged_student_ids:
        student = students_dict.get(sid)
        if not student:
            continue
        courses_data = []
        for card in student_cards_map[sid]:
            course = db.query(Course).get(card.course_id)
            class_obj = db.query(Class).get(card.class_id) if card.class_id else None
            total_leave_quota = card.total_leave_quota or 0
            used_leave = card.used_leave or 0
            if total_leave_quota == 0:
                remaining_leaves = 0
                leave_display = '不允许'
            else:
                remaining_leaves = total_leave_quota - used_leave
                leave_display = f"{used_leave}/{total_leave_quota}"
            validity_display = '-'
            if card.validity_end_date:
                validity_display = card.validity_end_date.strftime('%Y-%m-%d')
            elif card.manual_validity_days:
                validity_display = (date.today() + timedelta(days=card.manual_validity_days)).strftime('%Y-%m-%d')
            courses_data.append({
                "course_id": card.course_id,
                "course_name": course.name if course else '',
                "class_name": class_obj.name if class_obj else '未分班',
                "validity_display": validity_display,
                "remaining_hours": card.remaining_paid_hours,
                "remaining_amount": card.remaining_amount,
                "remaining_gift": card.remaining_gift_hours,
                "remaining_leaves": remaining_leaves,
                "leave_display": leave_display,
                "is_suspended": card.status == 'suspended'
            })
        result.append({
            "id": student.id,
            "name": student.name,
            "phone": student.phone,
            "gender": student.gender,
            "age": student.age,
            "avatar": ImageService.get_url(student.avatar),
            "card_background": ImageService.get_url(student.card_background),
            "total_integral": student.total_integral,
            "courses": courses_data
        })

    return {"code": 0, "data": {"items": result, "total": total}}


# ===================================================================
# 9. 学员作品管理
# ===================================================================

# 作品存储目录（相对于 media 根目录）
WORKS_ORIGINALS_DIR = "media/student_works/originals"
WORKS_THUMBNAILS_DIR = "media/student_works/thumbnails"
os.makedirs(WORKS_ORIGINALS_DIR, exist_ok=True)
os.makedirs(WORKS_THUMBNAILS_DIR, exist_ok=True)


@router.post("/student/works/{student_id}/upload")
async def upload_student_work(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """上传学员作品（图片或视频，最大100MB）"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")

    # 校验文件大小
    file.file.seek(0, 2)
    size = file.file.tell()
    file.file.seek(0)
    if size > 100 * 1024 * 1024:
        raise HTTPException(400, "文件大小不能超过 100MB")

    # 校验类型
    content_type = file.content_type or ""
    is_image = content_type.startswith("image/")
    is_video = content_type.startswith("video/")
    if not (is_image or is_video):
        raise HTTPException(400, "仅支持图片或视频文件")

    # 生成唯一文件名
    ext = os.path.splitext(file.filename)[1]
    if not ext:
        ext = ".jpg" if is_image else ".mp4"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    original_path = os.path.join(WORKS_ORIGINALS_DIR, unique_name)
    thumbnail_path = os.path.join(WORKS_THUMBNAILS_DIR, unique_name)

    # 保存原文件
    with open(original_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # 生成缩略图（仅图片）
    thumbnail_rel = None
    if is_image:
        try:
            img = PILImage.open(original_path)
            img.thumbnail((200, 200), PILImage.Resampling.LANCZOS)
            img.save(thumbnail_path, quality=85, optimize=True)
            thumbnail_rel = f"student_works/thumbnails/{unique_name}"
        except Exception as e:
            logger.warning(f"生成缩略图失败: {e}")
            thumbnail_rel = f"student_works/originals/{unique_name}"
    else:
        # 视频使用原图作为缩略图（前端可识别 type 后显示播放按钮）
        thumbnail_rel = f"student_works/originals/{unique_name}"

    # 保存到数据库
    work = StudentWork(
        student_id=student_id,
        name=file.filename,
        type="image" if is_image else "video",
        original_url=f"student_works/originals/{unique_name}",
        thumbnail_url=thumbnail_rel,
        size=size
    )
    db.add(work)
    db.commit()
    db.refresh(work)

    return {
        "code": 0,
        "data": {
            "id": work.id,
            "name": work.name,
            "type": work.type,
            "original_url": ImageService.get_url(work.original_url),
            "thumbnail_url": ImageService.get_url(work.thumbnail_url),
            "created_at": work.created_at.isoformat()
        },
        "message": "上传成功"
    }


@router.get("/student/works/{student_id}")
def get_student_works(
    student_id: int,
    db: Session = Depends(get_db)
):
    """获取学员作品列表（按时间倒序）"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")

    works = db.query(StudentWork).filter(
        StudentWork.student_id == student_id
    ).order_by(StudentWork.created_at.desc()).all()

    result = []
    for w in works:
        result.append({
            "id": w.id,
            "name": w.name,
            "type": w.type,
            "original_url": ImageService.get_url(w.original_url),
            "thumbnail_url": ImageService.get_url(w.thumbnail_url),
            "size": w.size,
            "created_at": w.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })
    return {"code": 0, "data": result}


@router.delete("/student/works/{student_id}")
def delete_student_works(
    student_id: int,
    data: DeleteWorksRequest,
    db: Session = Depends(get_db)
):
    """批量删除学员作品（同时删除物理文件）"""
    student = db.query(Student).get(student_id)
    if not student:
        raise HTTPException(404, "学员不存在")

    works = db.query(StudentWork).filter(
        StudentWork.student_id == student_id,
        StudentWork.id.in_(data.ids)
    ).all()

    if not works:
        raise HTTPException(404, "未找到指定作品")

    for w in works:
        # 删除物理文件
        original_path = os.path.join("media", w.original_url)
        if os.path.exists(original_path):
            try:
                os.remove(original_path)
            except OSError:
                pass
        if w.thumbnail_url:
            thumb_path = os.path.join("media", w.thumbnail_url)
            if os.path.exists(thumb_path) and thumb_path != original_path:
                try:
                    os.remove(thumb_path)
                except OSError:
                    pass
        db.delete(w)

    db.commit()
    return {"code": 0, "message": f"已删除 {len(works)} 个作品"}