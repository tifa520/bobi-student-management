from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request, Body
from sqlalchemy.orm import Session, joinedload
from typing import Optional, List
from datetime import datetime, timedelta
from decimal import Decimal, getcontext
import random
import io
import json
import openpyxl
from urllib.parse import quote
from loguru import logger

from app.database import get_db
from app.models import (
    Activity, ActivityRegistration, Student, MiscFee, IntegralRecord, SystemConfig,
    User, PrizeWinner, LearningCard, Attendance, Order, Course, Class
)
from app.schemas import ActivityRegister
from app.utils import get_current_user
from app.services.image_service import ImageService
from app.config import POINTS_TO_CASH_RATE

# 设置 Decimal 精度
getcontext().prec = 10

router = APIRouter()

# ==================== 工具函数 ====================

def get_current_exchange_rate(db: Session) -> int:
    """获取当前积分汇率（1元 = ? 积分）"""
    config = db.query(SystemConfig).filter(SystemConfig.key == 'exchange_rate').first()
    return int(config.value) if config else POINTS_TO_CASH_RATE


def _calc_activity_status(activity: Activity) -> str:
    """根据活动时间和报名时间计算状态"""
    now = datetime.now()
    if activity.status == 'draft':
        return 'draft'
    if activity.status == 'cancelled':
        return 'cancelled'
    if activity.status == 'archived':
        return 'archived'
    if activity.registration_start and now < activity.registration_start:
        return 'pending'
    if activity.registration_end and now > activity.registration_end:
        if activity.start_date and now < activity.start_date:
            return 'upcoming'
        elif activity.start_date and now < activity.end_date:
            return 'ongoing'
        elif activity.end_date and now > activity.end_date:
            return 'ended'
        else:
            return 'registering'
    else:
        if activity.start_date and now < activity.start_date:
            return 'registering'
        elif activity.start_date and now < activity.end_date:
            return 'ongoing'
        elif activity.end_date and now > activity.end_date:
            return 'ended'
        else:
            return 'registering'


def _update_activity_status(activity: Activity, db: Session):
    if activity.status in ['draft', 'cancelled', 'archived']:
        return
    new_status = _calc_activity_status(activity)
    if new_status != activity.status:
        activity.status = new_status
        db.add(activity)
        db.flush()


# ==================== 活动 CRUD ====================

@router.get("/activities")
def get_activities(
    request: Request,
    limit: int = Query(0, ge=0),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None),
    activity_type: Optional[str] = Query(None),
    is_featured: Optional[bool] = Query(None),
    is_archived: Optional[bool] = Query(False),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(Activity).filter(Activity.is_archived == (is_archived if is_archived is not None else False))
    if keyword:
        query = query.filter(Activity.name.contains(keyword))
    if status:
        query = query.filter(Activity.status == status)
    if activity_type:
        query = query.filter(Activity.activity_type == activity_type)
    if is_featured is not None:
        query = query.filter(Activity.is_featured == is_featured)
    
    activities = query.all()
    for a in activities:
        _update_activity_status(a, db)
    db.commit()
    
    query = query.order_by(Activity.is_featured.desc(), Activity.featured_order.asc(), Activity.created_at.desc())
    total = query.count()
    if limit > 0:
        paginated = query.limit(limit).all()
    else:
        paginated = query.offset((page - 1) * page_size).limit(page_size).all()
    
    result = []
    for a in paginated:
        reg_count = db.query(ActivityRegistration).filter(ActivityRegistration.activity_id == a.id).count()
        stats = a.stats or {}
        stats['registrations'] = reg_count
        stats['confirmed'] = db.query(ActivityRegistration).filter(
            ActivityRegistration.activity_id == a.id,
            ActivityRegistration.is_attending == '是'
        ).count()
        stats['paid'] = db.query(ActivityRegistration).filter(
            ActivityRegistration.activity_id == a.id,
            ActivityRegistration.is_paid == True
        ).count()
        result.append({
            "id": a.id,
            "name": a.name,
            "start_date": a.start_date.strftime("%Y-%m-%d %H:%M:%S") if a.start_date else '',
            "end_date": a.end_date.strftime("%Y-%m-%d %H:%M:%S") if a.end_date else '',
            "registration_start": a.registration_start.strftime("%Y-%m-%d %H:%M:%S") if a.registration_start else '',
            "registration_end": a.registration_end.strftime("%Y-%m-%d %H:%M:%S") if a.registration_end else '',
            "activity_type": a.activity_type or '其他',
            "charge_mode": a.charge_mode or 'free',
            "fee": a.fee,
            "points_cost": a.points_cost,
            "is_free": a.is_free,
            "content": a.content,
            "cover_image": ImageService.get_url(a.cover_image),
            "banner_image": ImageService.get_url(a.banner_image),
            "is_featured": a.is_featured,
            "featured_order": a.featured_order,
            "prizes": a.prizes or [],
            "stats": stats,
            "recommend_type": a.recommend_type,
            "pay_option": a.pay_option,
            "location": a.location,
            "max_participants": a.max_participants,
            "status": a.status,
            "is_archived": a.is_archived,
            "enable_lottery": a.enable_lottery or False,
            "lottery_times": a.lottery_times or 1,
            "max_win_times": a.max_win_times or 1,
            "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else '',
        })
    return {"code": 0, "data": {"items": result, "total": total}}


@router.get("/activities/{activity_id}")
def get_activity_detail(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    _update_activity_status(a, db)
    db.commit()
    
    reg_count = db.query(ActivityRegistration).filter(ActivityRegistration.activity_id == activity_id).count()
    stats = a.stats or {}
    stats['registrations'] = reg_count
    stats['confirmed'] = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.is_attending == '是'
    ).count()
    stats['paid'] = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.is_paid == True
    ).count()
    
    registrations = db.query(ActivityRegistration).filter(ActivityRegistration.activity_id == activity_id).all()
    reg_list = []
    for reg in registrations:
        student = db.query(Student).get(reg.student_id)
        reg_list.append({
            "id": reg.id,
            "student_id": reg.student_id,
            "student_name": student.name if student else '',
            "student_phone": student.phone if student else '',
            "student_avatar": ImageService.get_url(student.avatar) if student else '',
            "is_attending": reg.is_attending,
            "is_paid": reg.is_paid,
            "paid_amount": reg.paid_amount,
            "points_used": reg.points_used,
            "points_refunded": reg.points_refunded or 0,
            "refund_amount": reg.refund_amount,
            "status": reg.status,
            "lottery_count": reg.lottery_count or 0,
            "win_count": reg.win_count or 0,
            "created_at": reg.created_at.strftime("%Y-%m-%d %H:%M:%S") if reg.created_at else '',
            "cash_payment_method": reg.cash_payment_method or '',
        })
    
    return {
        "code": 0,
        "data": {
            "activity": {
                "id": a.id,
                "name": a.name,
                "start_date": a.start_date.strftime("%Y-%m-%d %H:%M:%S") if a.start_date else '',
                "end_date": a.end_date.strftime("%Y-%m-%d %H:%M:%S") if a.end_date else '',
                "registration_start": a.registration_start.strftime("%Y-%m-%d %H:%M:%S") if a.registration_start else '',
                "registration_end": a.registration_end.strftime("%Y-%m-%d %H:%M:%S") if a.registration_end else '',
                "activity_type": a.activity_type or '其他',
                "charge_mode": a.charge_mode or 'free',
                "fee": a.fee,
                "points_cost": a.points_cost,
                "is_free": a.is_free,
                "content": a.content,
                "cover_image": ImageService.get_url(a.cover_image),
                "banner_image": ImageService.get_url(a.banner_image),
                "is_featured": a.is_featured,
                "featured_order": a.featured_order,
                "prizes": a.prizes or [],
                "stats": stats,
                "recommend_type": a.recommend_type,
                "pay_option": a.pay_option,
                "location": a.location,
                "max_participants": a.max_participants,
                "status": a.status,
                "is_archived": a.is_archived,
                "enable_lottery": a.enable_lottery or False,
                "lottery_times": a.lottery_times or 1,
                "max_win_times": a.max_win_times or 1,
                "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else '',
            },
            "statistics": {
                "total_reg": len(reg_list),
                "confirmed": sum(1 for r in reg_list if r['is_attending'] == '是'),
                "paid_count": sum(1 for r in reg_list if r['is_paid']),
                "total_paid": sum(r['paid_amount'] for r in reg_list if r['is_paid']),
                "total_points_used": sum(r['points_used'] for r in reg_list),
                "pending": sum(1 for r in reg_list if r['is_attending'] == '待定'),
                "unpaid_confirmed": sum(1 for r in reg_list if r['is_attending'] == '是' and not r['is_paid'])
            },
            "registrations": reg_list
        }
    }


@router.post("/activities")
def create_activity(data: dict, request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user(request, db)
    except:
        current_user = None
    
    start_date = None
    end_date = None
    registration_start = None
    registration_end = None
    if data.get('start_date'):
        try:
            start_date = datetime.strptime(data['start_date'], "%Y-%m-%d %H:%M:%S")
        except:
            start_date = datetime.strptime(data['start_date'], "%Y-%m-%d")
    if data.get('end_date'):
        try:
            end_date = datetime.strptime(data['end_date'], "%Y-%m-%d %H:%M:%S")
        except:
            end_date = datetime.strptime(data['end_date'], "%Y-%m-%d")
    if data.get('registration_start'):
        try:
            registration_start = datetime.strptime(data['registration_start'], "%Y-%m-%d %H:%M:%S")
        except:
            registration_start = datetime.strptime(data['registration_start'], "%Y-%m-%d")
    if data.get('registration_end'):
        try:
            registration_end = datetime.strptime(data['registration_end'], "%Y-%m-%d %H:%M:%S")
        except:
            registration_end = datetime.strptime(data['registration_end'], "%Y-%m-%d")
    
    a = Activity(
        name=data.get('name'),
        start_date=start_date,
        end_date=end_date,
        registration_start=registration_start,
        registration_end=registration_end,
        activity_type=data.get('activity_type', '其他'),
        charge_mode=data.get('charge_mode', 'free'),
        is_free=data.get('is_free', True),
        fee=data.get('fee', 0.0),
        points_cost=data.get('points_cost', 0),
        content=data.get('content', ''),
        cover_image=data.get('cover_image'),
        banner_image=data.get('banner_image'),
        is_featured=data.get('is_featured', False),
        featured_order=data.get('featured_order', 0),
        prizes=data.get('prizes', []),
        recommend_type=data.get('recommend_type', 'grid'),
        pay_option=data.get('pay_option', 'both'),
        location=data.get('location', ''),
        max_participants=data.get('max_participants', 0),
        status='draft',
        is_archived=False,
        created_by=current_user.id if current_user else None,
        enable_lottery=data.get('enable_lottery', False),
        lottery_times=data.get('lottery_times', 1),
        max_win_times=data.get('max_win_times', 1)
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return {"code": 0, "data": {"id": a.id}, "message": "创建成功"}


@router.put("/activities/{activity_id}")
def update_activity(activity_id: int, data: dict, request: Request, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    
    allowed_fields = [
        'name', 'start_date', 'end_date', 'registration_start', 'registration_end',
        'activity_type', 'charge_mode', 'is_free', 'fee', 'points_cost', 'content',
        'cover_image', 'banner_image', 'is_featured', 'featured_order', 'prizes',
        'recommend_type', 'pay_option', 'location', 'max_participants', 'status', 'is_archived',
        'enable_lottery', 'lottery_times', 'max_win_times'
    ]
    for field in allowed_fields:
        if field in data:
            if field in ['start_date', 'end_date', 'registration_start', 'registration_end'] and data[field]:
                try:
                    setattr(a, field, datetime.strptime(data[field], "%Y-%m-%d %H:%M:%S"))
                except:
                    setattr(a, field, datetime.strptime(data[field], "%Y-%m-%d"))
            else:
                setattr(a, field, data[field])
    
    db.commit()
    return {"code": 0, "message": "更新成功"}


@router.delete("/activities/{activity_id}")
def delete_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    if db.query(ActivityRegistration).filter(ActivityRegistration.activity_id == activity_id).count() > 0:
        raise HTTPException(400, "该活动已有学员报名，无法删除")
    db.delete(a)
    db.commit()
    return {"code": 0, "message": "已删除"}


@router.post("/activities/{activity_id}/publish")
def publish_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    if a.status != 'draft':
        raise HTTPException(400, "只有草稿状态的活动可以发布")
    a.status = 'registering'
    db.commit()
    return {"code": 0, "message": "活动已发布"}


@router.post("/activities/{activity_id}/cancel")
def cancel_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    if a.status in ['cancelled', 'archived']:
        raise HTTPException(400, "活动已取消或归档，无法再次取消")
    
    registrations = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.is_paid == True
    ).all()
    for reg in registrations:
        if reg.paid_amount > 0:
            refund_misc = MiscFee(
                student_id=reg.student_id,
                fee_type='refund',
                source_id=reg.id,
                description=f"活动【{a.name}】取消退款 {reg.paid_amount} 元",
                amount=-reg.paid_amount,
                paid_amount=reg.paid_amount,
                status='paid',
                payment_method=reg.cash_payment_method or '原路返回',
                paid_at=datetime.now()
            )
            db.add(refund_misc)
            reg.refund_amount = (reg.refund_amount or 0) + reg.paid_amount
            reg.paid_amount = 0
            reg.is_paid = False
        if reg.points_used > 0:
            student = db.query(Student).get(reg.student_id)
            if student:
                student.total_integral += reg.points_used
                integral_record = IntegralRecord(
                    student_id=reg.student_id,
                    change_amount=reg.points_used,
                    reason=f"活动【{a.name}】取消返还积分"
                )
                db.add(integral_record)
                reg.points_refunded = (reg.points_refunded or 0) + reg.points_used
                reg.points_used = 0
        reg.refund_at = datetime.now()
        reg.status = 'cancelled'
    
    a.status = 'cancelled'
    db.commit()
    return {"code": 0, "message": f"已取消活动，共处理 {len(registrations)} 笔退款"}


@router.post("/activities/{activity_id}/archive")
def archive_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    if a.status != 'ended':
        raise HTTPException(400, "只有已结束的活动可以归档")
    a.is_archived = True
    db.commit()
    return {"code": 0, "message": "已归档"}


@router.post("/activities/{activity_id}/unarchive")
def unarchive_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, "活动不存在")
    a.is_archived = False
    db.commit()
    return {"code": 0, "message": "已取消归档"}


# ==================== 活动图片上传 ====================

@router.post("/upload-image")
async def upload_activity_image(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        get_current_user(request, db)
    except:
        raise HTTPException(401, "未授权")
    
    relative_path = await ImageService.save(file, "activity_image")
    full_url = ImageService.get_url(relative_path)
    return {"code": 0, "data": {"url": full_url}, "message": "上传成功"}


# ==================== 报名管理 ====================

@router.post("/activities/{activity_id}/register")
def register_activity(
    activity_id: int,
    data: ActivityRegister,
    request: Request,
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    if activity.status in ['cancelled', 'archived']:
        raise HTTPException(400, "活动已取消或归档，无法报名")
    
    student = db.query(Student).get(data.student_id)
    if not student:
        raise HTTPException(404, "学员不存在")

    existing = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.student_id == data.student_id,
        ActivityRegistration.status != 'cancelled'
    ).first()
    if existing:
        raise HTTPException(400, "该学员已报名此活动")

    if activity.max_participants > 0:
        current_count = db.query(ActivityRegistration).filter(
            ActivityRegistration.activity_id == activity_id
        ).count()
        if current_count >= activity.max_participants:
            raise HTTPException(400, "活动报名人数已达上限")

    # 计算应付金额
    if activity.charge_mode == 'free':
        total_fee = Decimal('0')
    elif activity.charge_mode == 'paid':
        total_fee = Decimal(str(activity.fee))
    else:
        total_fee = Decimal('0')

    exchange_rate = get_current_exchange_rate(db)

    # 解析支付方式
    payments = data.payments or []
    if not payments and total_fee == 0:
        payments = []
    elif not payments:
        raise HTTPException(400, "请选择支付方式")

    total_cash = Decimal('0')
    total_points = Decimal('0')
    payment_details = []

    for p in payments:
        if p.get('type') == 'cash':
            amount = Decimal(str(p.get('amount', 0)))
            if amount < 0:
                raise HTTPException(400, "现金金额不能为负数")
            total_cash += amount
            payment_details.append({
                "type": "cash",
                "method": p.get('method', '现金'),
                "amount": float(amount)
            })
        elif p.get('type') == 'points':
            points = Decimal(str(p.get('points', 0)))
            if points < 0:
                raise HTTPException(400, "积分不能为负数")
            total_points += points
            payment_details.append({
                "type": "points",
                "method": "积分",
                "points": int(points)
            })
        else:
            raise HTTPException(400, "不支持的支付类型")

    # ★ 使用 Decimal 精确校验支付总额
    paid_value = total_cash + total_points / Decimal(str(exchange_rate))
    if paid_value != total_fee:
        raise HTTPException(400, f"支付总额 {float(paid_value):.2f} 与应付金额 {float(total_fee):.2f} 不匹配")

    # 校验积分余额
    if total_points > Decimal(str(student.total_integral)):
        raise HTTPException(400, f"积分不足，需要 {int(total_points)} 积分，当前 {student.total_integral}")

    # 扣减积分
    points_used = int(total_points)
    if points_used > 0:
        student.total_integral -= points_used
        integral_record = IntegralRecord(
            student_id=student.id,
            change_amount=-points_used,
            reason=f"报名活动【{activity.name}】抵扣 {points_used/exchange_rate:.2f} 元"
        )
        db.add(integral_record)

    # 创建报名记录
    is_paid = (total_fee == 0) or (paid_value == total_fee)

    reg = ActivityRegistration(
        activity_id=activity_id,
        student_id=data.student_id,
        is_attending=data.is_attending,
        is_paid=is_paid,
        paid_amount=float(total_cash),
        points_used=points_used,
        refund_amount=0,
        points_refunded=0,
        cash_payment_method=payment_details[0].get('method', '') if payment_details and payment_details[0]['type'] == 'cash' else '',
        status='confirmed',
        operator_id=None,
        lottery_count=0,
        win_count=0
    )
    db.add(reg)
    db.flush()

    # 记录杂费（收费活动）
    if total_fee > 0:
        misc = MiscFee(
            student_id=student.id,
            fee_type='activity',
            source_id=reg.id,
            description=f"参加活动【{activity.name}】",
            amount=float(total_fee),
            paid_amount=float(total_cash),
            points_used=points_used,
            exchange_rate=exchange_rate,
            status='paid' if is_paid else 'partial',
            payment_method=json.dumps(payment_details, ensure_ascii=False) if payment_details else '',
            paid_at=datetime.now() if is_paid else None
        )
        db.add(misc)

    stats = activity.stats or {}
    stats['registrations'] = (stats.get('registrations', 0)) + 1
    activity.stats = stats
    db.add(activity)

    db.commit()
    return {"code": 0, "data": {"id": reg.id}, "message": "报名成功"}


@router.get("/activities/{activity_id}/registrations")
def get_activity_registrations(activity_id: int, db: Session = Depends(get_db)):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    registrations = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id
    ).all()
    result = []
    for reg in registrations:
        student = db.query(Student).get(reg.student_id)
        result.append({
            "id": reg.id,
            "student_id": reg.student_id,
            "student_name": student.name if student else '',
            "student_phone": student.phone if student else '',
            "student_avatar": ImageService.get_url(student.avatar) if student else '',
            "is_attending": reg.is_attending,
            "is_paid": reg.is_paid,
            "paid_amount": reg.paid_amount,
            "points_used": reg.points_used,
            "points_refunded": reg.points_refunded or 0,
            "refund_amount": reg.refund_amount,
            "status": reg.status,
            "lottery_count": reg.lottery_count or 0,
            "win_count": reg.win_count or 0,
            "created_at": reg.created_at.strftime("%Y-%m-%d %H:%M:%S") if reg.created_at else '',
            "cash_payment_method": reg.cash_payment_method or '',
        })
    return {"code": 0, "data": result}


@router.put("/registrations/{registration_id}")
def update_registration(
    registration_id: int,
    data: dict,
    db: Session = Depends(get_db)
):
    reg = db.query(ActivityRegistration).get(registration_id)
    if not reg:
        raise HTTPException(404, "报名记录不存在")
    
    allowed = ['is_attending', 'is_paid', 'paid_amount', 'points_used', 'cash_payment_method']
    for field in allowed:
        if field in data:
            setattr(reg, field, data[field])
    reg.updated_at = datetime.now()
    db.commit()
    return {"code": 0, "message": "更新成功"}


@router.post("/registrations/{registration_id}/cancel")
def cancel_registration(
    registration_id: int,
    reason: str = Query(''),
    db: Session = Depends(get_db)
):
    # ★ 加行锁防止并发
    reg = db.query(ActivityRegistration).filter(
        ActivityRegistration.id == registration_id
    ).with_for_update().first()
    
    if not reg:
        raise HTTPException(404, "报名记录不存在")
    
    # ★ 幂等检查1：如果已取消，直接返回
    if reg.status == 'cancelled':
        return {"code": 0, "message": "报名已取消，无需重复操作"}
    
    # ★ 幂等检查2：如果已经退过费，不再重复退费
    if reg.refund_amount > 0 or reg.points_refunded > 0:
        reg.status = 'cancelled'
        db.commit()
        return {"code": 0, "message": "报名已取消（费用已退还）"}
    
    activity = db.query(Activity).get(reg.activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    # 计算可退金额和积分
    refund_amount = reg.paid_amount or 0
    refund_points = (reg.points_used or 0) - (reg.points_refunded or 0)
    
    # 退现金
    if refund_amount > 0:
        refund_misc = MiscFee(
            student_id=reg.student_id,
            fee_type='refund',
            source_id=reg.id,
            description=f"活动【{activity.name}】取消报名退款 {refund_amount} 元",
            amount=-refund_amount,
            paid_amount=refund_amount,
            status='paid',
            payment_method=reg.cash_payment_method or '原路返回',
            paid_at=datetime.now()
        )
        db.add(refund_misc)
        reg.refund_amount = (reg.refund_amount or 0) + refund_amount
        reg.paid_amount = 0
        reg.is_paid = False
    
    # 退积分
    if refund_points > 0:
        student = db.query(Student).get(reg.student_id)
        if student:
            student.total_integral += refund_points
            integral_record = IntegralRecord(
                student_id=reg.student_id,
                change_amount=refund_points,
                reason=f"活动【{activity.name}】取消报名返还积分"
            )
            db.add(integral_record)
            reg.points_refunded = (reg.points_refunded or 0) + refund_points
            reg.points_used = (reg.points_used or 0) - refund_points
    
    reg.refund_at = datetime.now()
    reg.status = 'cancelled'
    reg.cancel_reason = reason
    db.commit()
    return {"code": 0, "message": "取消报名成功"}


@router.delete("/registrations/{registration_id}")
def delete_registration(registration_id: int, db: Session = Depends(get_db)):
    reg = db.query(ActivityRegistration).get(registration_id)
    if not reg:
        raise HTTPException(404, "报名记录不存在")
    if reg.is_paid or reg.paid_amount > 0 or reg.points_used > 0:
        raise HTTPException(400, "该报名已有缴费记录，请先退费再删除")
    db.delete(reg)
    db.commit()
    return {"code": 0, "message": "已删除"}


@router.post("/activities/{activity_id}/batch-pay")
def batch_pay_registrations(
    activity_id: int,
    registration_ids: List[int] = Body(...),
    is_paid: bool = Query(...),
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    db.query(ActivityRegistration).filter(
        ActivityRegistration.id.in_(registration_ids),
        ActivityRegistration.activity_id == activity_id
    ).update({"is_paid": is_paid}, synchronize_session=False)
    db.commit()
    return {"code": 0, "message": "批量更新成功"}


@router.post("/registrations/{registration_id}/refund")
def refund_registration(
    registration_id: int,
    refund_amount: float = Query(0),
    refund_points: int = Query(0),
    payment_method: str = Query('原路返回'),
    db: Session = Depends(get_db)
):
    reg = db.query(ActivityRegistration).get(registration_id)
    if not reg:
        raise HTTPException(404, "报名记录不存在")
    
    activity = db.query(Activity).get(reg.activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")

    max_cash_refund = reg.paid_amount
    if refund_amount > max_cash_refund:
        raise HTTPException(400, f"现金退款不能超过实付金额 {max_cash_refund}")

    max_points_refund = reg.points_used - (reg.points_refunded or 0)
    if refund_points > max_points_refund:
        raise HTTPException(400, f"积分退款不能超过已用积分 {max_points_refund}")

    student = db.query(Student).get(reg.student_id)
    if not student:
        raise HTTPException(404, "学员不存在")

    if refund_amount > 0:
        refund_misc = MiscFee(
            student_id=reg.student_id,
            fee_type='refund',
            source_id=reg.id,
            description=f"活动【{activity.name}】退款 {refund_amount} 元",
            amount=-refund_amount,
            paid_amount=refund_amount,
            status='paid',
            payment_method=reg.cash_payment_method or payment_method,
            paid_at=datetime.now()
        )
        db.add(refund_misc)
        reg.refund_amount = (reg.refund_amount or 0) + refund_amount
        reg.paid_amount -= refund_amount
        if reg.paid_amount <= 0:
            reg.is_paid = False

    if refund_points > 0:
        student.total_integral += refund_points
        integral_record = IntegralRecord(
            student_id=reg.student_id,
            change_amount=refund_points,
            reason=f"活动【{activity.name}】退费返还积分"
        )
        db.add(integral_record)
        reg.points_refunded = (reg.points_refunded or 0) + refund_points
        reg.points_used -= refund_points

    reg.refund_at = datetime.now()
    db.commit()
    return {"code": 0, "message": f"退费成功，现金 {refund_amount} 元，积分 {refund_points} 分"}


# ★★★ ==================== 奖项管理 ====================

@router.put("/activities/{activity_id}/prizes")
def set_activity_prizes(activity_id: int, data: List[dict], db: Session = Depends(get_db)):
    """保存活动的奖项设置（直接写入活动.prizes字段）"""
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    activity.prizes = data
    db.commit()
    return {"code": 0, "message": "奖项设置已保存"}


# ★★★ ==================== 抽奖功能（含行锁防并发） ====================

@router.get("/activities/{activity_id}/lottery/status")
def get_lottery_status(
    activity_id: int,
    student_id: int = Query(...),
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    if not activity.enable_lottery:
        raise HTTPException(400, "该活动未开启抽奖")
    
    reg = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.student_id == student_id
    ).first()
    if not reg:
        raise HTTPException(404, "学员未报名此活动")
    
    return {
        "code": 0,
        "data": {
            "remaining_draws": max(0, activity.lottery_times - reg.lottery_count),
            "remaining_wins": max(0, activity.max_win_times - reg.win_count),
            "can_draw": reg.lottery_count < activity.lottery_times,
            "can_win": reg.win_count < activity.max_win_times,
            "total_draws": activity.lottery_times,
            "total_wins": activity.max_win_times,
            "prizes": [p for p in (activity.prizes or []) if p.get('remaining', 0) > 0]
        }
    }


@router.post("/activities/{activity_id}/lottery/draw")
def draw_lottery(
    activity_id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """
    执行抽奖 - 使用数据库行锁（FOR UPDATE）保证并发安全
    """
    student_id = data.get('student_id')
    if not student_id:
        raise HTTPException(400, "缺少student_id")

    # ★ 1. 锁定活动行（FOR UPDATE）
    activity = db.query(Activity).filter(
        Activity.id == activity_id
    ).with_for_update().first()
    if not activity:
        raise HTTPException(404, "活动不存在")
    if not activity.enable_lottery:
        raise HTTPException(400, "该活动未开启抽奖")

    # ★ 2. 锁定报名记录
    reg = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id,
        ActivityRegistration.student_id == student_id
    ).with_for_update().first()
    if not reg:
        raise HTTPException(404, "学员未报名此活动")

    # ★ 3. 检查剩余抽奖次数（已在锁保护下）
    if reg.lottery_count >= activity.lottery_times:
        raise HTTPException(400, "抽奖次数已用完")

    # ★ 4. 处理奖项池（深拷贝防止回滚失败）
    prizes = [dict(p) for p in (activity.prizes or [])]
    available_prizes = [p for p in prizes if p.get('remaining', 0) > 0]

    win_result = {"win": False, "prize_name": None, "prize_level": None}

    # ★ 5. 判断是否可中奖
    if reg.win_count < activity.max_win_times and available_prizes:
        # 按剩余数量加权随机抽取
        weights = [p['remaining'] for p in available_prizes]
        chosen = random.choices(available_prizes, weights=weights, k=1)[0]

        # 扣减库存（在 prizes 副本上操作）
        for p in prizes:
            if p['name'] == chosen['name'] and p.get('level') == chosen.get('level'):
                p['remaining'] -= 1
                break

        # 创建中奖记录
        winner = PrizeWinner(
            activity_id=activity_id,
            registration_id=reg.id,
            prize_id=None,
            prize_name=chosen['name'],
            prize_level=chosen.get('level', '其他'),
            status='pending'
        )
        db.add(winner)
        reg.win_count += 1
        win_result = {
            "win": True,
            "prize_name": chosen['name'],
            "prize_level": chosen.get('level', '其他')
        }

    # ★ 6. 更新活动奖项JSON & 抽奖次数
    activity.prizes = prizes
    reg.lottery_count += 1

    # ★ 7. 提交事务（释放行锁）
    db.commit()

    return {
        "code": 0,
        "data": {
            **win_result,
            "remaining_draws": activity.lottery_times - reg.lottery_count,
            "remaining_wins": activity.max_win_times - reg.win_count
        }
    }


# ★★★ ==================== 中奖名单管理 ====================

@router.get("/activities/{activity_id}/winners")
def get_winners(activity_id: int, db: Session = Depends(get_db)):
    winners = db.query(PrizeWinner).filter(PrizeWinner.activity_id == activity_id).all()
    result = []
    for w in winners:
        reg = db.query(ActivityRegistration).get(w.registration_id)
        student = db.query(Student).get(reg.student_id) if reg else None
        result.append({
            "id": w.id,
            "student_id": student.id if student else None,
            "student_name": student.name if student else '',
            "student_avatar": ImageService.get_url(student.avatar) if student else '',
            "prize_name": w.prize_name,
            "prize_level": w.prize_level,
            "status": w.status,
            "created_at": w.created_at.strftime("%Y-%m-%d %H:%M:%S") if w.created_at else '',
        })
    return {"code": 0, "data": result}


@router.put("/winners/{winner_id}")
def update_winner(winner_id: int, data: dict, db: Session = Depends(get_db)):
    w = db.query(PrizeWinner).get(winner_id)
    if not w:
        raise HTTPException(404, "中奖记录不存在")
    allowed = ['status']
    for field in allowed:
        if field in data:
            setattr(w, field, data[field])
    if data.get('status') in ['delivered', 'claimed']:
        w.delivered_at = datetime.now()
    db.commit()
    return {"code": 0, "message": "更新成功"}


@router.delete("/winners/{winner_id}")
def delete_winner(winner_id: int, db: Session = Depends(get_db)):
    w = db.query(PrizeWinner).get(winner_id)
    if not w:
        raise HTTPException(404, "中奖记录不存在")
    db.delete(w)
    db.commit()
    return {"code": 0, "message": "已删除"}


# ==================== 导入导出 ====================

@router.get("/activities/{activity_id}/registrations/export")
def export_registrations(
    activity_id: int,
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    registrations = db.query(ActivityRegistration).filter(
        ActivityRegistration.activity_id == activity_id
    ).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报名名单"
    headers = ["学员姓名", "手机号", "是否参加", "缴费状态", "实付金额", "使用积分", "退款金额", "退款积分", "支付方式", "报名时间"]
    ws.append(headers)
    for reg in registrations:
        student = db.query(Student).get(reg.student_id)
        ws.append([
            student.name if student else '',
            student.phone if student else '',
            reg.is_attending,
            '已缴费' if reg.is_paid else '未缴费',
            reg.paid_amount,
            reg.points_used,
            reg.refund_amount,
            reg.points_refunded or 0,
            reg.cash_payment_method or '',
            reg.created_at.strftime("%Y-%m-%d %H:%M:%S") if reg.created_at else ''
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"报名名单_{activity.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.post("/activities/{activity_id}/registrations/import")
async def import_registrations(
    activity_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value.strip()] = idx
    
    required = ['学员姓名', '手机号']
    for field in required:
        if field not in headers:
            return {"code": 1, "message": f"模板缺少必填字段：{field}"}
    
    imported = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[headers.get('学员姓名', 0)-1]:
            continue
        try:
            name = str(row[headers['学员姓名']-1]).strip()
            phone = str(row[headers['手机号']-1]).strip()
            student = db.query(Student).filter(Student.phone == phone).first()
            if not student:
                student = Student(
                    name=name,
                    phone=phone,
                    student_uuid=str(uuid.uuid4())
                )
                db.add(student)
                db.flush()
            
            existing = db.query(ActivityRegistration).filter(
                ActivityRegistration.activity_id == activity_id,
                ActivityRegistration.student_id == student.id
            ).first()
            if existing:
                errors.append(f"第{row_idx}行：学员 {name} 已报名")
                continue
            
            is_attending = str(row[headers.get('是否参加', 0)-1]).strip() if headers.get('是否参加') else '是'
            paid_amount = float(row[headers.get('实付金额', 0)-1]) if headers.get('实付金额') and row[headers.get('实付金额', 0)-1] else 0
            points_used = int(row[headers.get('使用积分', 0)-1]) if headers.get('使用积分') and row[headers.get('使用积分', 0)-1] else 0
            payment_method = str(row[headers.get('支付方式', 0)-1]).strip() if headers.get('支付方式') else ''
            
            reg = ActivityRegistration(
                activity_id=activity_id,
                student_id=student.id,
                is_attending=is_attending,
                is_paid=paid_amount > 0 or points_used > 0,
                paid_amount=paid_amount,
                points_used=points_used,
                cash_payment_method=payment_method,
                status='confirmed',
                lottery_count=0,
                win_count=0
            )
            db.add(reg)
            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")
    
    db.commit()
    return {"code": 0, "message": f"成功导入 {imported} 条记录，{len(errors)} 条错误", "data": {"imported": imported, "errors": errors[:10]}}


@router.post("/activities/{activity_id}/winners/import")
async def import_winners(
    activity_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value.strip()] = idx
    
    required = ['学员姓名', '手机号', '奖品名称']
    for field in required:
        if field not in headers:
            return {"code": 1, "message": f"模板缺少必填字段：{field}"}
    
    imported = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[headers.get('学员姓名', 0)-1]:
            continue
        try:
            name = str(row[headers['学员姓名']-1]).strip()
            phone = str(row[headers['手机号']-1]).strip()
            prize_name = str(row[headers['奖品名称']-1]).strip()
            prize_level = str(row[headers.get('奖品等级', 0)-1]).strip() if headers.get('奖品等级') else '其他'
            
            student = db.query(Student).filter(Student.phone == phone).first()
            if not student:
                errors.append(f"第{row_idx}行：学员 {name} 不存在")
                continue
            
            reg = db.query(ActivityRegistration).filter(
                ActivityRegistration.activity_id == activity_id,
                ActivityRegistration.student_id == student.id
            ).first()
            if not reg:
                errors.append(f"第{row_idx}行：学员 {name} 未报名此活动")
                continue
            
            existing = db.query(PrizeWinner).filter(
                PrizeWinner.activity_id == activity_id,
                PrizeWinner.registration_id == reg.id,
                PrizeWinner.prize_name == prize_name
            ).first()
            if existing:
                errors.append(f"第{row_idx}行：学员 {name} 已中奖 {prize_name}")
                continue
            
            winner = PrizeWinner(
                activity_id=activity_id,
                registration_id=reg.id,
                prize_id=None,
                prize_name=prize_name,
                prize_level=prize_level or '其他',
                status='pending'
            )
            db.add(winner)
            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")
    
    db.commit()
    return {
        "code": 0,
        "message": f"成功导入 {imported} 条中奖记录，{len(errors)} 条错误",
        "data": {"imported": imported, "errors": errors[:10]}
    }


@router.get("/activities/{activity_id}/winners/export")
def export_winners(
    activity_id: int,
    db: Session = Depends(get_db)
):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    winners = db.query(PrizeWinner).filter(PrizeWinner.activity_id == activity_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "中奖名单"
    headers = ["学员姓名", "手机号", "奖品名称", "奖项等级", "状态", "中奖时间"]
    ws.append(headers)
    for w in winners:
        reg = db.query(ActivityRegistration).get(w.registration_id)
        student = db.query(Student).get(reg.student_id) if reg else None
        status_map = {'pending': '待发放', 'delivered': '已发放', 'claimed': '已领取'}
        ws.append([
            student.name if student else '',
            student.phone if student else '',
            w.prize_name,
            w.prize_level,
            status_map.get(w.status, w.status),
            w.created_at.strftime("%Y-%m-%d %H:%M:%S") if w.created_at else ''
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"中奖名单_{activity.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ==================== 统计接口 ====================

@router.get("/activities/{activity_id}/stats")
def get_activity_stats(activity_id: int, db: Session = Depends(get_db)):
    activity = db.query(Activity).get(activity_id)
    if not activity:
        raise HTTPException(404, "活动不存在")
    
    regs = db.query(ActivityRegistration).filter(ActivityRegistration.activity_id == activity_id).all()
    total_reg = len(regs)
    confirmed = sum(1 for r in regs if r.is_attending == '是')
    paid = sum(1 for r in regs if r.is_paid)
    attended = sum(1 for r in regs if r.status == 'attended')
    total_paid_amount = sum(r.paid_amount for r in regs)
    total_points_used = sum(r.points_used for r in regs)
    total_refund_amount = sum(r.refund_amount for r in regs)
    total_refund_points = sum(r.points_refunded or 0 for r in regs)

    # 计算奖品总成本（从活动.prizes字段）
    prizes = activity.prizes or []
    total_prize_cost = sum(p.get('cost', 0) * p.get('quantity', 0) for p in prizes)

    net_revenue = total_paid_amount - total_refund_amount - total_prize_cost

    return {
        "code": 0,
        "data": {
            "total_registrations": total_reg,
            "confirmed": confirmed,
            "paid": paid,
            "attended": attended,
            "total_paid_amount": total_paid_amount,
            "total_points_used": total_points_used,
            "total_refund_amount": total_refund_amount,
            "total_refund_points": total_refund_points,
            "total_prize_cost": total_prize_cost,
            "net_revenue": net_revenue,
            "conversion_rate": round(confirmed / total_reg * 100, 2) if total_reg > 0 else 0,
            "attendance_rate": round(attended / confirmed * 100, 2) if confirmed > 0 else 0,
            "payments": [
                {
                    "student_name": db.query(Student).get(r.student_id).name if db.query(Student).get(r.student_id) else '',
                    "student_avatar": ImageService.get_url(db.query(Student).get(r.student_id).avatar) if db.query(Student).get(r.student_id) else '',
                    "paid_amount": r.paid_amount,
                    "points_used": r.points_used,
                    "refund_amount": r.refund_amount,
                    "points_refunded": r.points_refunded or 0,
                    "cash_payment_method": r.cash_payment_method or ''
                }
                for r in regs if r.paid_amount > 0 or r.points_used > 0
            ]
        }
    }